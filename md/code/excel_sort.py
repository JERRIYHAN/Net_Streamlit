from openpyxl import *

def read_excel():
    #第1步：读取顺序表，该表指定了排序key的顺序，在下面2行填入表格路径、工作表名称
    #若无现成顺序表，这一段亦可以不写。直接用new列表指定排序key顺序，可设计由用户输入append
    wb1=load_workbook('file_path')
    sheet1= wb1['sheet_name']
    new=[]
    for j in range(3,18):#在表中指定位置寻找排序列，此处是3-17列
        cell1 = sheet1.cell(row=1,column=j) #在表中指定位置寻找排序行，此处是第1行
        new.append(cell1.value)
        
    print(new) #new列表即为key顺序

    #第2步：批量读取待排序表格
    for i in range(24,31):#取决于表格个数，以i命名文件————下一行
        wb = load_workbook('file_path'+str(i))
        sheet = wb['Sheet1']
        # print(sheet)
        
    #第3步：读取具体数据，输出为列表col_lst
        col_lst = []
        for col in sheet.columns: #这里是按行排序，若按列排序，可将col换为row，columns换为rows
            col_data = [cell.value for cell in col]
            col_lst.append(col_data)

        #print(col_lst)

    #第4步：排序：根据顺序append进新列表output，此处太繁琐待优化
        output=[]
        for k in range(15): 
            for p in range(15):
                if col_lst[p][0]==new[k]:
                    output.append(col_lst[p])
        #print(output)
                    
    #第5步：将output输出到新表格
        wb2= Workbook()
        s_sheet = wb2.create_sheet("排序")       # 创建新的sheet
        for jj in range(len(output[0])):
            for ii in range(len(output)):
                s_sheet.cell(jj+1,ii+1,output[ii][jj]) # 如果需要行列转换就把ii+1，jj+1换一下

        #保存文件
        wb2.save('file_path'+str(i)+'Sort'+'.xlsx')                    

read_excel()
print('finish')
