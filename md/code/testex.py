from openpyxl import *

wb = Workbook()
ws = wb.create_sheet("Mysheet")

ws.cell(1,1,88)
wb.save('/Users/jiaaijiajiazhangshi/Downloads/test.xlsx')
